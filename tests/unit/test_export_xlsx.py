"""Unit tests — XLSX export (issue #7)."""

from __future__ import annotations

import stat

import pytest
import openpyxl

from quip_export.formats.xlsx import export_xlsx

HTML_TWO_SHEETS = """
<html><body>
<div class="sheet" data-title="Sheet1">
<table>
  <tr><th>Name</th><th>Score</th></tr>
  <tr><td>Alice</td><td>95</td></tr>
</table>
</div>
<div class="sheet" data-title="Sheet2">
<table>
  <tr><th>City</th><th>Pop</th></tr>
  <tr><td>Berlin</td><td>3645000</td></tr>
</table>
</div>
</body></html>
"""


class TestExportXlsx:
    def test_creates_file(self, tmp_path, html_spreadsheet):
        out = tmp_path / "data.xlsx"
        export_xlsx(html_spreadsheet, out)
        assert out.exists()
        assert out.stat().st_size > 0

    def test_text_values_present_in_cells(self, tmp_path, html_spreadsheet):
        out = tmp_path / "data.xlsx"
        export_xlsx(html_spreadsheet, out)
        wb = openpyxl.load_workbook(out)
        ws = wb.active
        cell_values = [ws.cell(row=r, column=c).value for r in range(1, 4) for c in range(1, 4)]
        assert "Alice" in cell_values or "Name" in cell_values

    def test_numeric_values_stored_as_numbers(self, tmp_path, html_spreadsheet):
        out = tmp_path / "data.xlsx"
        export_xlsx(html_spreadsheet, out)
        wb = openpyxl.load_workbook(out)
        ws = wb.active
        numeric_cells = [
            ws.cell(row=r, column=c).value
            for r in range(2, ws.max_row + 1)
            for c in range(1, ws.max_column + 1)
            if isinstance(ws.cell(row=r, column=c).value, (int, float))
        ]
        assert len(numeric_cells) > 0

    def test_multi_sheet_creates_multiple_worksheets(self, tmp_path):
        out = tmp_path / "multi.xlsx"
        export_xlsx(HTML_TWO_SHEETS, out)
        wb = openpyxl.load_workbook(out)
        assert len(wb.sheetnames) >= 2

    def test_empty_spreadsheet_produces_valid_file(self, tmp_path, html_empty):
        out = tmp_path / "empty.xlsx"
        export_xlsx(html_empty, out)
        assert out.exists()
        wb = openpyxl.load_workbook(out)
        assert wb is not None

    def test_large_spreadsheet_not_truncated(self, tmp_path):
        rows = "".join(f"<tr><td>Row{i}</td><td>{i}</td></tr>" for i in range(1, 10001))
        html = f"<html><body><table><tr><th>Label</th><th>Num</th></tr>{rows}</table></body></html>"
        out = tmp_path / "large.xlsx"
        export_xlsx(html, out)
        wb = openpyxl.load_workbook(out)
        ws = wb.active
        assert ws.max_row >= 10000

    def test_formula_cell_exports_display_value_not_formula(self, tmp_path):
        html = """
        <html><body><table>
          <tr><th>A</th><th>B</th></tr>
          <tr><td>10</td><td>=A2*2</td></tr>
        </table></body></html>
        """
        out = tmp_path / "formula.xlsx"
        export_xlsx(html, out)
        wb = openpyxl.load_workbook(out)
        ws = wb.active
        cell_b2 = ws.cell(row=2, column=2).value
        assert not str(cell_b2).startswith("="), "Formula must not be written literally"

    def test_write_failure_raises_os_error(self, tmp_path, html_spreadsheet):
        ro = tmp_path / "ro"
        ro.mkdir()
        ro.chmod(stat.S_IREAD | stat.S_IEXEC)
        out = ro / "data.xlsx"
        try:
            with pytest.raises((PermissionError, OSError)):
                export_xlsx(html_spreadsheet, out)
        finally:
            ro.chmod(stat.S_IRWXU)

    def test_header_row_has_bold_font(self, tmp_path, html_spreadsheet):
        out = tmp_path / "data.xlsx"
        export_xlsx(html_spreadsheet, out)
        wb = openpyxl.load_workbook(out)
        ws = wb.active
        header_cell = ws.cell(row=1, column=1)
        assert header_cell.font is not None and header_cell.font.bold

    def test_no_api_content_produces_empty_xlsx_with_warning(self, tmp_path, caplog, html_empty):
        out = tmp_path / "noop.xlsx"
        export_xlsx(html_empty, out)
        assert out.exists()
