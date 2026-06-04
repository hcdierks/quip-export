"""Export Quip spreadsheet HTML to XLSX."""

from __future__ import annotations

import logging
from pathlib import Path

import openpyxl
from bs4 import BeautifulSoup, Tag
from openpyxl.styles import Font

log = logging.getLogger(__name__)


def _cell_value(text: str) -> int | float | str:
    """Convert cell text to the most appropriate Python type.

    Formula-like strings (starting with '=') are stored as plain text
    to avoid being evaluated as Excel formulas.
    """
    if text.startswith("="):
        return "'" + text  # prefix forces Excel to treat as literal string
    try:
        return int(text)
    except ValueError:
        pass
    try:
        return float(text)
    except ValueError:
        pass
    return text


def _write_table(ws: openpyxl.worksheet.worksheet.Worksheet, table: Tag) -> None:
    rows = table.find_all("tr")
    for row_idx, row in enumerate(rows, start=1):
        cells = row.find_all(["th", "td"])
        is_header = bool(row.find("th"))
        for col_idx, cell in enumerate(cells, start=1):
            value = cell.get_text(strip=True)
            ws.cell(row=row_idx, column=col_idx, value=_cell_value(value))
            if is_header:
                ws.cell(row=row_idx, column=col_idx).font = Font(bold=True)


def export_xlsx(html: str, output_path: Path) -> None:
    soup = BeautifulSoup(html, "lxml")
    wb = openpyxl.Workbook()
    # Remove the auto-created default sheet
    wb.remove(wb.active)  # type: ignore[arg-type]

    sheet_divs = soup.find_all("div", class_="sheet")
    if sheet_divs:
        for div in sheet_divs:
            title = div.get("data-title", "Sheet") or "Sheet"
            ws = wb.create_sheet(title=str(title))
            table = div.find("table")
            if table:
                _write_table(ws, table)  # type: ignore[arg-type]
    else:
        ws = wb.create_sheet(title="Sheet1")
        table = soup.find("table")
        if table:
            _write_table(ws, table)  # type: ignore[arg-type]
        else:
            log.warning("No table found in HTML for XLSX export")

    if not wb.sheetnames:
        wb.create_sheet("Sheet1")

    wb.save(output_path)
